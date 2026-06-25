# 请帮我写个中文的 Python 脚本，批注也是中文，但是变量参数不要是中文：
# 询问我需要的操作：
# 1. 移动：询问我 excel 路径（默认"d:\Studios\Attachments\标准.xlsx"）。源文件夹路径（默认"d:\Studios\Folders\Ins\"）与目标文件夹路径（默认"d:\Studios\Folders\Outs\"）。读取 excel 文件，第一行为表头（字段名），此后每一行为一条记录。依次根据每一行的"原文件名"，在源文件夹路径找到后移动到目标文件夹。
# 2. 复制：询问我 excel 路径（默认"d:\Studios\Attachments\标准.xlsx"）。源文件夹路径（默认"d:\Studios\Folders\Ins\"）与目标文件夹路径（默认"d:\Studios\Folders\Outs\"）。读取 excel 文件，第一行为表头（字段名），此后每一行为一条记录。依次根据每一行的"原文件名"，在源文件夹路径找到后复制到目标文件夹。
# 3. 重命名：询问我 excel 路径（默认"d:\Studios\Attachments\标准.xlsx"）。源文件夹路径（默认"d:\Studios\Folders\Ins\"）。读取 excel 文件，第一行为表头（字段名），此后每一行为一条记录。依次根据每一行的"原文件名"，在源文件夹路径找到后改名为"现文件名"。
# 4. 记录写入：询问我 excel 路径（默认"d:\Studios\Attachments\标准.xlsx"）。源文件夹路径（默认"d:\Studios\Folders\Ins\"）。询问我是否包含子文件夹里的文件。读取 excel 文件，此后每一行为一条记录。根据源文件夹下文件写入该 excel 文件。要求只添加记录，不修改已经生成的记录。1. 每一条记录新开一行。2. "Index"字段值为上一行 "Index"字段值+1（如上一行为空或为表头，则"Index"字段值为 1）。3. 将源文件夹路径里的所有文件的文件名写入 excel 文件的"名字"与"原文件名"字段值"（名字字段不包括扩展名，原文件名包括扩展名）。
# 5. 记录删除：询问我 excel 路径（默认"d:\Studios\Attachments\标准.xlsx"）。源文件夹路径（默认"d:\Studios\Folders\Ins\"）。询问我是否包含子文件夹里的文件；询问我匹配的字段是"原文件名"还是"现文件名"。读取 excel 文件，此后每一行为一条记录。枚举源文件夹路径里的所有文件的文件名, 将匹配到 excel 文件的相应字段值的行（记录）进行删除。最后提示我在源文件夹里还有什么文件没有在原 excel 文件里，列出那些文件。
# 结束后重复询问。

import signal
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

# ==================== 全局配置 ====================

DEFAULT_EXCEL_PATH = Path(r"d:\Studios\Attachments\标准.xlsx")
DEFAULT_SOURCE_DIR = Path(r"d:\Studios\Folders\Ins")
DEFAULT_TARGET_DIR = Path(r"d:\Studios\Folders\Outs")

_quit_requested = False  # Ctrl+Q 中断标志

# --- 消息常量 ---
MSG_INTERRUPTED = "\n\n用户中断程序，已退出。"
MSG_ERROR = "\n程序运行出错: {}"
MSG_EXIT = "\n按回车键退出..."
MSG_ERR_EXCEL_NOT_FOUND = "错误：Excel 文件不存在 -> {}"
MSG_ERR_SRC_DIR_NOT_EXIST = "错误：源文件夹不存在 -> {}"
MSG_ERR_MISSING_COL = "错误：Excel 中缺少'{}'列。"
MSG_ERR_MISSING_ORIG_OR_NEW_COL = "错误：Excel 中缺少'原文件名'或'现文件名'列。"
MSG_ERR_MISSING_INDEX_NAME_ORIG_COL = "错误：Excel 表头必须包含'Index'、'名字'、'原文件名'列。"
MSG_WARN_FILE_NOT_FOUND = "警告：未找到文件，跳过 -> {}"
MSG_WARN_TARGET_EXISTS = "警告：目标已存在，跳过 -> {}"
MSG_MOVED = "已移动: {}"
MSG_MOVE_COMPLETE = "移动完成，共处理 {} 个文件。"
MSG_COPIED = "已复制: {}"
MSG_COPY_COMPLETE = "复制完成，共处理 {} 个文件。"
MSG_RENAMED = "重命名: {} -> {}"
MSG_RENAME_COMPLETE = "重命名完成，共处理 {} 个文件。"
MSG_EXCEL_NEW_FILE = "Excel 文件不存在，将创建新文件并添加表头 -> {}"
MSG_RECORDS_APPENDED = "已追加 {} 条记录到 Excel，保存至: {}"
MSG_DELETED_RECORDS = "已删除 {} 条匹配'{}'的记录（共 {} 行数据）。"
MSG_MISSING_FILES_HEADER = "\n源文件夹中以下文件未在 Excel 的'{}'列中出现（共 {} 个）："
MSG_MISSING_FILE_ITEM = "  - {}"
MSG_ALL_FILES_COVERED = "源文件夹中的所有文件均在 Excel 的'{}'列中有对应记录。"
MSG_TOOL_TITLE = "===== 文件管理工具（Excel 驱动）====="
MSG_OPT_MOVE = "1. 移动"
MSG_OPT_COPY = "2. 复制"
MSG_OPT_RENAME = "3. 重命名"
MSG_OPT_WRITE = "4. 记录写入"
MSG_OPT_DELETE = "5. 记录删除"
MSG_OPT_EXIT = "0. 退出"
MSG_ASK_CHOICE = "请输入数字(0-5): "
MSG_PROGRAM_EXIT = "程序已退出。"
MSG_ASK_EXCEL_PATH = "Excel 文件路径"
MSG_ASK_SRC_DIR = "源文件夹路径"
MSG_ASK_DST_DIR = "目标文件夹路径"
MSG_ASK_INCLUDE_SUBDIRS = "是否包含子文件夹？(y/n)"
MSG_ASK_MATCH_FIELD = "匹配的字段名（原文件名/现文件名）"
MSG_WARN_UNKNOWN_FIELD = "警告：未知字段'{}'，将使用默认的'原文件名'。"
MSG_INVALID_CHOICE = "无效的选择，请输入 0-5 之间的数字。"
MSG_PRESS_ENTER = "\n按回车键继续..."
MSG_SECTION_MOVE = "\n--- 移动文件 ---"
MSG_SECTION_COPY = "\n--- 复制文件 ---"
MSG_SECTION_RENAME = "\n--- 重命名文件 ---"
MSG_SECTION_WRITE = "\n--- 记录写入 ---"
MSG_SECTION_DELETE = "\n--- 记录删除 ---"
MSG_DEFAULT_PROMPT_SUFFIX = " (默认: {}): "
COL_ORIG_NAME = "原文件名"
COL_NEW_NAME = "现文件名"

# ==================== 辅助函数 ====================


def get_input_with_default(prompt_text: str, default_value: str) -> str:
    """获取用户输入，直接回车则使用默认值。"""
    user_input = input(f"{prompt_text}{MSG_DEFAULT_PROMPT_SUFFIX.format(default_value)}").strip()
    return user_input if user_input else str(default_value)


def get_column_index(ws, col_name: str) -> int | None:
    """在表头第一行中查找列名，返回 1-based 列索引；未找到返回 None。"""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip() == col_name:
            return cell.column
    return None


def get_last_index_value(ws, index_col: int) -> int:
    """获取 Index 列最后一行的值。无数据行返回 0。"""
    for row in range(ws.max_row, 1, -1):
        cell = ws.cell(row=row, column=index_col)
        if cell.value is not None:
            return cell.value
    return 0


# ==================== 核心操作 ====================


def move_files_from_excel(excel_path: Path, src_dir: Path, dst_dir: Path) -> None:
    """根据 Excel 中的'原文件名'列移动文件。"""
    if not excel_path.is_file():
        print(MSG_ERR_EXCEL_NOT_FOUND.format(excel_path))
        return
    if not src_dir.is_dir():
        print(MSG_ERR_SRC_DIR_NOT_EXIST.format(src_dir))
        return
    dst_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(str(excel_path))
    ws = wb.active
    col = get_column_index(ws, COL_ORIG_NAME)
    if col is None:
        print(MSG_ERR_MISSING_COL.format(COL_ORIG_NAME))
        wb.close()
        return

    moved = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, values_only=False):
        name = row[col - 1].value
        if not name:
            continue
        name = str(name).strip()
        src_path = src_dir / name
        dst_path = dst_dir / name
        if not src_path.is_file():
            print(MSG_WARN_FILE_NOT_FOUND.format(src_path))
            continue
        if dst_path.exists():
            print(MSG_WARN_TARGET_EXISTS.format(dst_path))
            continue
        shutil.move(str(src_path), str(dst_path))
        print(MSG_MOVED.format(name))
        moved += 1

    wb.close()
    print(MSG_MOVE_COMPLETE.format(moved))


def copy_files_from_excel(excel_path: Path, src_dir: Path, dst_dir: Path) -> None:
    """根据 Excel 中的'原文件名'列复制文件。"""
    if not excel_path.is_file():
        print(MSG_ERR_EXCEL_NOT_FOUND.format(excel_path))
        return
    if not src_dir.is_dir():
        print(MSG_ERR_SRC_DIR_NOT_EXIST.format(src_dir))
        return
    dst_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(str(excel_path))
    ws = wb.active
    col = get_column_index(ws, COL_ORIG_NAME)
    if col is None:
        print(MSG_ERR_MISSING_COL.format(COL_ORIG_NAME))
        wb.close()
        return

    copied = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, values_only=False):
        name = row[col - 1].value
        if not name:
            continue
        name = str(name).strip()
        src_path = src_dir / name
        dst_path = dst_dir / name
        if not src_path.is_file():
            print(MSG_WARN_FILE_NOT_FOUND.format(src_path))
            continue
        if dst_path.exists():
            print(MSG_WARN_TARGET_EXISTS.format(dst_path))
            continue
        shutil.copy2(str(src_path), str(dst_path))
        print(MSG_COPIED.format(name))
        copied += 1

    wb.close()
    print(MSG_COPY_COMPLETE.format(copied))


def rename_files_from_excel(excel_path: Path, src_dir: Path) -> None:
    """根据 Excel 中的'原文件名'和'现文件名'列重命名文件。"""
    if not excel_path.is_file():
        print(MSG_ERR_EXCEL_NOT_FOUND.format(excel_path))
        return
    if not src_dir.is_dir():
        print(MSG_ERR_SRC_DIR_NOT_EXIST.format(src_dir))
        return

    wb = load_workbook(str(excel_path))
    ws = wb.active
    col_old = get_column_index(ws, COL_ORIG_NAME)
    col_new = get_column_index(ws, COL_NEW_NAME)
    if col_old is None or col_new is None:
        print(MSG_ERR_MISSING_ORIG_OR_NEW_COL)
        wb.close()
        return

    renamed = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, values_only=False):
        old_name = row[col_old - 1].value
        new_name = row[col_new - 1].value
        if not old_name or not new_name:
            continue
        old_name = str(old_name).strip()
        new_name = str(new_name).strip()
        old_path = src_dir / old_name
        new_path = src_dir / new_name
        if not old_path.is_file():
            print(MSG_WARN_FILE_NOT_FOUND.format(old_path))
            continue
        if new_path.exists():
            print(MSG_WARN_TARGET_EXISTS.format(new_path))
            continue
        old_path.rename(new_path)
        print(MSG_RENAMED.format(old_name, new_name))
        renamed += 1

    wb.close()
    print(MSG_RENAME_COMPLETE.format(renamed))


def write_filenames_to_excel(excel_path: Path, src_dir: Path, include_subdirs: bool) -> None:
    """将源文件夹中的文件名追加写入 Excel。"""
    if not src_dir.is_dir():
        print(MSG_ERR_SRC_DIR_NOT_EXIST.format(src_dir))
        return

    if not excel_path.is_file():
        print(MSG_EXCEL_NEW_FILE.format(excel_path))
        wb = Workbook()
        ws = wb.active
        ws.append(["Index", "名字", "原文件名"])
        wb.save(str(excel_path))
        wb.close()

    wb = load_workbook(str(excel_path))
    ws = wb.active
    max_col = ws.max_column

    col_index = get_column_index(ws, "Index")
    col_name = get_column_index(ws, "名字")
    col_orig = get_column_index(ws, COL_ORIG_NAME)
    if col_index is None or col_name is None or col_orig is None:
        print(MSG_ERR_MISSING_INDEX_NAME_ORIG_COL)
        wb.close()
        return

    last_index = get_last_index_value(ws, col_index)
    next_index = last_index + 1

    # 收集文件
    iterator = src_dir.rglob("*") if include_subdirs else src_dir.iterdir()
    file_paths = [p for p in iterator if p.is_file()]

    added = 0
    for f in file_paths:
        new_row = [""] * max_col
        new_row[col_index - 1] = next_index
        new_row[col_name - 1] = f.stem
        new_row[col_orig - 1] = f.name
        ws.append(new_row)
        added += 1
        next_index += 1

    wb.save(str(excel_path))
    wb.close()
    print(MSG_RECORDS_APPENDED.format(added, excel_path))


def delete_records_from_excel(
    excel_path: Path, src_dir: Path, include_subdirs: bool, match_field: str = COL_ORIG_NAME,
) -> None:
    """删除 Excel 中匹配源文件夹文件名的行，报告未匹配文件。"""
    if not excel_path.is_file():
        print(MSG_ERR_EXCEL_NOT_FOUND.format(excel_path))
        return
    if not src_dir.is_dir():
        print(MSG_ERR_SRC_DIR_NOT_EXIST.format(src_dir))
        return

    # 收集源文件夹文件名
    iterator = src_dir.rglob("*") if include_subdirs else src_dir.iterdir()
    src_names = {p.name for p in iterator if p.is_file()}

    wb = load_workbook(str(excel_path))
    ws = wb.active

    col_match = get_column_index(ws, match_field)
    if col_match is None:
        print(MSG_ERR_MISSING_COL.format(match_field))
        wb.close()
        return

    header = [cell.value for cell in ws[1]]
    rows_to_keep = [header]
    excel_names: set[str] = set()
    deleted = 0
    total_rows = 0

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, values_only=True):
        total_rows += 1
        val = row[col_match - 1] if col_match - 1 < len(row) else None
        val = str(val).strip() if val else ""
        if val:
            excel_names.add(val)
        if val in src_names:
            deleted += 1
            continue
        rows_to_keep.append(list(row))

    ws.delete_rows(1, ws.max_row)
    for row_data in rows_to_keep:
        ws.append(row_data)
    wb.save(str(excel_path))
    wb.close()

    missing = src_names - excel_names
    print(MSG_DELETED_RECORDS.format(deleted, match_field, total_rows))
    if missing:
        print(MSG_MISSING_FILES_HEADER.format(match_field, len(missing)))
        for f in sorted(missing):
            print(MSG_MISSING_FILE_ITEM.format(f))
    else:
        print(MSG_ALL_FILES_COVERED.format(match_field))


# ==================== 主程序 ====================
# ==================== 中断处理 ====================


def _on_quit_signal(signum, frame):
    global _quit_requested
    _quit_requested = True
    raise KeyboardInterrupt()


def _init_quit_handler():
    if hasattr(signal, "SIGQUIT"):
        signal.signal(signal.SIGQUIT, _on_quit_signal)


def _check_quit() -> bool:
    global _quit_requested
    if sys.platform == "win32":
        try:
            import msvcrt
            while msvcrt.kbhit():
                if msvcrt.getch() == b"\x11":
                    _quit_requested = True
        except Exception:
            pass
    return _quit_requested


def main() -> None:
    while True:
        print("\n" + MSG_TOOL_TITLE)
        print(MSG_OPT_MOVE)
        print(MSG_OPT_COPY)
        print(MSG_OPT_RENAME)
        print(MSG_OPT_WRITE)
        print(MSG_OPT_DELETE)
        print(MSG_OPT_EXIT)
        choice = input(MSG_ASK_CHOICE).strip()

        if choice == "0":
            print(MSG_PROGRAM_EXIT)
            break

        elif choice == "1":
            print(MSG_SECTION_MOVE)
            excel = Path(get_input_with_default(MSG_ASK_EXCEL_PATH, str(DEFAULT_EXCEL_PATH)))
            src = Path(get_input_with_default(MSG_ASK_SRC_DIR, str(DEFAULT_SOURCE_DIR)))
            dst = Path(get_input_with_default(MSG_ASK_DST_DIR, str(DEFAULT_TARGET_DIR)))
            move_files_from_excel(excel, src, dst)

        elif choice == "2":
            print(MSG_SECTION_COPY)
            excel = Path(get_input_with_default(MSG_ASK_EXCEL_PATH, str(DEFAULT_EXCEL_PATH)))
            src = Path(get_input_with_default(MSG_ASK_SRC_DIR, str(DEFAULT_SOURCE_DIR)))
            dst = Path(get_input_with_default(MSG_ASK_DST_DIR, str(DEFAULT_TARGET_DIR)))
            copy_files_from_excel(excel, src, dst)

        elif choice == "3":
            print(MSG_SECTION_RENAME)
            excel = Path(get_input_with_default(MSG_ASK_EXCEL_PATH, str(DEFAULT_EXCEL_PATH)))
            src = Path(get_input_with_default(MSG_ASK_SRC_DIR, str(DEFAULT_SOURCE_DIR)))
            rename_files_from_excel(excel, src)

        elif choice == "4":
            print(MSG_SECTION_WRITE)
            excel = Path(get_input_with_default(MSG_ASK_EXCEL_PATH, str(DEFAULT_EXCEL_PATH)))
            src = Path(get_input_with_default(MSG_ASK_SRC_DIR, str(DEFAULT_SOURCE_DIR)))
            inc = get_input_with_default(MSG_ASK_INCLUDE_SUBDIRS, "n").strip().lower() == "y"
            write_filenames_to_excel(excel, src, inc)

        elif choice == "5":
            print(MSG_SECTION_DELETE)
            excel = Path(get_input_with_default(MSG_ASK_EXCEL_PATH, str(DEFAULT_EXCEL_PATH)))
            src = Path(get_input_with_default(MSG_ASK_SRC_DIR, str(DEFAULT_SOURCE_DIR)))
            inc = get_input_with_default(MSG_ASK_INCLUDE_SUBDIRS, "n").strip().lower() == "y"
            field = get_input_with_default(MSG_ASK_MATCH_FIELD, COL_ORIG_NAME)
            if field not in (COL_ORIG_NAME, COL_NEW_NAME):
                print(MSG_WARN_UNKNOWN_FIELD.format(field))
                field = COL_ORIG_NAME
            delete_records_from_excel(excel, src, inc, field)

        else:
            print(MSG_INVALID_CHOICE)

        input(MSG_PRESS_ENTER)


# ==================== 程序入口 ====================

if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    try:
        main()
    except KeyboardInterrupt:
        print(MSG_INTERRUPTED)
    except Exception as e:
        print(MSG_ERROR.format(e))
    finally:
        input(MSG_EXIT)
