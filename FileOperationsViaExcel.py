# 请帮我写个中文的 Python 脚本，批注也是中文，但是变量参数不要是中文：
# 询问我需要的操作：1. 移动； 2. 复制；3. 重命名；4. 记录写入；5. 记录删除。

# 1. 移动：询问我 excel 路径（默认“d:\\Works\\Attachments\\标准.xlsx”）。源文件夹路径（默认“d:\\Works\\Ins\\”）与目标文件夹路径（默认“d:\\Works\\Outs\\”）。读取 excel 文件，第一行为表头（字段名），此后每一行为一条记录。依次根据每一行的“原文件名”，在源文件夹路径找到后移动到目标文件夹。
# 2. 复制：询问我 excel 路径（默认“d:\\Works\\Attachments\\标准.xlsx”）。源文件夹路径（默认“d:\\Works\\Ins\\”）与目标文件夹路径（默认“d:\\Works\\Outs\\”）。读取 excel 文件，第一行为表头（字段名），此后每一行为一条记录。依次根据每一行的“原文件名”，在源文件夹路径找到后复制到目标文件夹。
# 3. 重命名：询问我 excel 路径（默认“d:\\Works\\Attachments\\标准.xlsx”）。源文件夹路径（默认“d:\\Works\\Ins\\”）。读取 excel 文件，第一行为表头（字段名），此后每一行为一条记录。依次根据每一行的“原文件名”，在源文件夹路径找到后改名为“现文件名”。
# 4. 记录写入：询问我 excel 路径（默认“d:\\Works\\Attachments\\标准.xlsx”）。源文件夹路径（默认“d:\\Works\\Ins\\”）。询问我是否包含子文件夹里的文件。读取 excel 文件，此后每一行为一条记录。根据源文件夹下文件写入该 excel 文件。要求只添加记录，不修改已经生成的记录。1. 每一条记录新开一行。2. "Index"字段值为上一行 "Index"字段值+1（如上一行为空或为表头，则"Index"字段值为 1）。3. 将源文件夹路径里的所有文件的文件名写入 excel 文件的"名字"与"原文件名"字段值”（名字字段不包括扩展名，原文件名包括扩展名）。
# 5. 记录删除：询问我 excel 路径（默认“d:\\Works\\Attachments\\标准.xlsx”）。源文件夹路径（默认“d:\\Works\\Ins\\”）。询问我是否包含子文件夹里的文件；询问我匹配的字段是“原文件名”还是“现文件名”。读取 excel 文件，此后每一行为一条记录。枚举源文件夹路径里的所有文件的文件名, 将匹配到 excel 文件的相应字段值的行（记录）进行删除。最后提示我在源文件夹里还有什么文件没有在原 excel 文件里，列出那些文件。
# 结束后重复询问。

# 导入模块
import os
import shutil
import sys
from openpyxl import load_workbook, Workbook

# --------------------------------------------
# 辅助函数
# --------------------------------------------

def get_input(prompt, default):
    """获取用户输入，直接回车则使用默认值"""
    user_input = input(f"{prompt} (默认: {default}): ").strip()
    return user_input if user_input else default

def get_column_index(ws, col_name):
    """
    在表头（第一行）中查找列名，返回列索引（1-based）。
    如果找不到，返回 None。
    """
    for cell in ws[1]:
        if cell.value and str(cell.value).strip() == col_name:
            return cell.column
    return None

def get_last_index_value(ws, index_col):
    """
    获取 Index 列最后一行的值。
    从最后一行向上查找，直到找到非空单元格或到达第2行。
    返回 (last_row, value)，如果没有任何数据行，返回 (1, 0)
    """
    last_row = ws.max_row
    while last_row >= 2:
        cell = ws.cell(row=last_row, column=index_col)
        if cell.value is not None:
            return last_row, cell.value
        last_row -= 1
    return 1, 0  # 没有数据行，Index 从 0 开始计数，加1后为1

# --------------------------------------------
# 核心操作函数
# --------------------------------------------

def move_files_from_excel(excel_path, src_dir, dst_dir):
    """根据 Excel 中的‘原文件名’列移动文件"""
    if not os.path.isfile(excel_path):
        print(f"错误：Excel 文件不存在 -> {excel_path}")
        return
    if not os.path.isdir(src_dir):
        print(f"错误：源文件夹不存在 -> {src_dir}")
        return
    os.makedirs(dst_dir, exist_ok=True)

    wb = load_workbook(excel_path)
    ws = wb.active

    # 查找所需的列
    col_original = get_column_index(ws, "原文件名")
    if col_original is None:
        print("错误：Excel 中缺少‘原文件名’列。")
        wb.close()
        return

    moved_count = 0
    # 从第2行开始遍历数据行
    for row in ws.iter_rows(min_row=2, values_only=False):
        original_name = row[col_original - 1].value  # row 是 0-based
        if not original_name:
            continue
        original_name = str(original_name).strip()
        src_path = os.path.join(src_dir, original_name)
        dst_path = os.path.join(dst_dir, original_name)
        if not os.path.isfile(src_path):
            print(f"警告：未找到文件，跳过 -> {src_path}")
            continue
        if os.path.exists(dst_path):
            print(f"警告：目标已存在，跳过 -> {dst_path}")
            continue
        shutil.move(src_path, dst_path)
        print(f"已移动: {original_name}")
        moved_count += 1

    wb.close()
    print(f"移动完成，共处理 {moved_count} 个文件。")

def copy_files_from_excel(excel_path, src_dir, dst_dir):
    """根据 Excel 中的‘原文件名’列复制文件"""
    if not os.path.isfile(excel_path):
        print(f"错误：Excel 文件不存在 -> {excel_path}")
        return
    if not os.path.isdir(src_dir):
        print(f"错误：源文件夹不存在 -> {src_dir}")
        return
    os.makedirs(dst_dir, exist_ok=True)

    wb = load_workbook(excel_path)
    ws = wb.active

    col_original = get_column_index(ws, "原文件名")
    if col_original is None:
        print("错误：Excel 中缺少‘原文件名’列。")
        wb.close()
        return

    copied_count = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        original_name = row[col_original - 1].value
        if not original_name:
            continue
        original_name = str(original_name).strip()
        src_path = os.path.join(src_dir, original_name)
        dst_path = os.path.join(dst_dir, original_name)
        if not os.path.isfile(src_path):
            print(f"警告：未找到文件，跳过 -> {src_path}")
            continue
        if os.path.exists(dst_path):
            print(f"警告：目标已存在，跳过 -> {dst_path}")
            continue
        shutil.copy2(src_path, dst_path)
        print(f"已复制: {original_name}")
        copied_count += 1

    wb.close()
    print(f"复制完成，共处理 {copied_count} 个文件。")

def rename_files_from_excel(excel_path, src_dir):
    """根据 Excel 中的‘原文件名’和‘现文件名’列重命名文件"""
    if not os.path.isfile(excel_path):
        print(f"错误：Excel 文件不存在 -> {excel_path}")
        return
    if not os.path.isdir(src_dir):
        print(f"错误：源文件夹不存在 -> {src_dir}")
        return

    wb = load_workbook(excel_path)
    ws = wb.active

    col_original = get_column_index(ws, "原文件名")
    col_new = get_column_index(ws, "现文件名")
    if col_original is None or col_new is None:
        print("错误：Excel 中缺少‘原文件名’或‘现文件名’列。")
        wb.close()
        return

    renamed_count = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        original_name = row[col_original - 1].value
        new_name = row[col_new - 1].value
        if not original_name or not new_name:
            continue
        original_name = str(original_name).strip()
        new_name = str(new_name).strip()
        old_path = os.path.join(src_dir, original_name)
        new_path = os.path.join(src_dir, new_name)
        if not os.path.isfile(old_path):
            print(f"警告：未找到文件，跳过 -> {old_path}")
            continue
        if os.path.exists(new_path):
            print(f"警告：目标名称已存在，跳过 -> {new_path}")
            continue
        os.rename(old_path, new_path)
        print(f"重命名: {original_name} -> {new_name}")
        renamed_count += 1

    wb.close()
    print(f"重命名完成，共处理 {renamed_count} 个文件。")

def write_filenames_to_excel(excel_path, src_dir, include_subdirs):
    """将源文件夹中的文件名（不含/含扩展名）追加写入 Excel"""
    if not os.path.isdir(src_dir):
        print(f"错误：源文件夹不存在 -> {src_dir}")
        return

    # 如果 Excel 文件不存在，则创建新文件并写入表头
    if not os.path.isfile(excel_path):
        print(f"Excel 文件不存在，将创建新文件并添加表头 -> {excel_path}")
        wb = Workbook()
        ws = wb.active
        ws.append(["Index", "名字", "原文件名"])
        wb.save(excel_path)
        wb.close()

    wb = load_workbook(excel_path)
    ws = wb.active

    # 获取表头总列数，用于构建新行
    max_col = ws.max_column
    # 确定关键列的位置
    col_index = get_column_index(ws, "Index")
    col_name = get_column_index(ws, "名字")
    col_original = get_column_index(ws, "原文件名")
    if col_index is None or col_name is None or col_original is None:
        print("错误：Excel 表头必须包含‘Index’、‘名字’、‘原文件名’列。")
        wb.close()
        return

    # 获取最后一行 Index 值，以便递增
    last_data_row, last_index = get_last_index_value(ws, col_index)
    next_index = last_index + 1

    # 收集源文件夹中的文件
    file_paths = []
    if include_subdirs:
        for root, dirs, files in os.walk(src_dir):
            for file in files:
                file_paths.append(os.path.join(root, file))
    else:
        for entry in os.listdir(src_dir):
            full_path = os.path.join(src_dir, entry)
            if os.path.isfile(full_path):
                file_paths.append(full_path)

    added_count = 0
    for full_path in file_paths:
        file_name = os.path.basename(full_path)                # 含扩展名
        base_name = os.path.splitext(file_name)[0]             # 不含扩展名

        # 构造新行列表，长度等于表头列数，非关键列留空
        new_row = [""] * max_col
        new_row[col_index - 1] = next_index
        new_row[col_name - 1] = base_name
        new_row[col_original - 1] = file_name

        ws.append(new_row)
        added_count += 1
        next_index += 1

    wb.save(excel_path)
    wb.close()
    print(f"已追加 {added_count} 条记录到 Excel，保存至: {excel_path}")

def delete_records_from_excel(excel_path, src_dir, include_subdirs, match_field="原文件名"):
    """
    删除 Excel 中指定字段值与源文件夹内实际文件名匹配的行，
    并报告源文件夹中有哪些文件在 Excel 中未出现（基于该字段）。
    """
    if not os.path.isfile(excel_path):
        print(f"错误：Excel 文件不存在 -> {excel_path}")
        return
    if not os.path.isdir(src_dir):
        print(f"错误：源文件夹不存在 -> {src_dir}")
        return

    # 收集源文件夹中的文件名（含扩展名）
    src_files = []
    if include_subdirs:
        for root, dirs, files in os.walk(src_dir):
            for file in files:
                src_files.append(file)
    else:
        for entry in os.listdir(src_dir):
            full_path = os.path.join(src_dir, entry)
            if os.path.isfile(full_path):
                src_files.append(entry)

    src_files_set = set(src_files)

    wb = load_workbook(excel_path)
    ws = wb.active

    col_match = get_column_index(ws, match_field)
    if col_match is None:
        print(f"错误：Excel 中缺少‘{match_field}’列。")
        wb.close()
        return

    # 读取所有行（包含表头），筛选出需要保留的行
    rows_to_keep = []
    # 先保留表头
    header = [cell.value for cell in ws[1]]
    rows_to_keep.append(header)

    # 收集 Excel 中匹配字段出现过的所有值（用于缺失报告）
    excel_files_set = set()
    deleted_count = 0
    total_data_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_data_rows += 1
        match_value = row[col_match - 1] if col_match - 1 < len(row) else None
        if match_value is not None:
            match_value = str(match_value).strip()
        else:
            match_value = ""
        # 空值不视为有效文件名，但为了统计仍加入集合
        if match_value:
            excel_files_set.add(match_value)
        # 如果该值存在于源文件夹文件名集合中，则删除该行
        if match_value in src_files_set:
            deleted_count += 1
            continue
        # 否则保留该行
        rows_to_keep.append(list(row))

    # 清空工作表并重新写入保留的行
    ws.delete_rows(1, ws.max_row)
    for row_data in rows_to_keep:
        ws.append(row_data)

    wb.save(excel_path)
    wb.close()

    # 找出源文件夹中有但 Excel 中没有的文件（基于匹配字段值）
    missing_files = src_files_set - excel_files_set

    print(f"已删除 {deleted_count} 条匹配‘{match_field}’的记录（共 {total_data_rows} 行数据）。")
    if missing_files:
        print(f"\n源文件夹中以下文件未在 Excel 的‘{match_field}’列中出现（共 {len(missing_files)} 个）：")
        for f in sorted(missing_files):
            print(f"  - {f}")
    else:
        print(f"源文件夹中的所有文件均在 Excel 的‘{match_field}’列中有对应记录。")

# --------------------------------------------
# 主程序
# --------------------------------------------

def main():
    default_excel = "d:\\Works\\Attachments\\标准.xlsx"
    default_src   = "d:\\Works\\Ins\\"
    default_dst   = "d:\\Works\\Outs\\"

    while True:
        print("\n===== 文件管理工具（Excel 驱动）=====")
        print("1. 移动")
        print("2. 复制")
        print("3. 重命名")
        print("4. 记录写入")
        print("5. 记录删除")
        print("0. 退出")
        choice = input("请输入数字(0-5): ").strip()

        if choice == '0':
            print("程序已退出。")
            break

        elif choice == '1':
            print("\n--- 移动文件 ---")
            excel_path = get_input("Excel 文件路径", default_excel)
            src_dir    = get_input("源文件夹路径", default_src)
            dst_dir    = get_input("目标文件夹路径", default_dst)
            move_files_from_excel(excel_path, src_dir, dst_dir)

        elif choice == '2':
            print("\n--- 复制文件 ---")
            excel_path = get_input("Excel 文件路径", default_excel)
            src_dir    = get_input("源文件夹路径", default_src)
            dst_dir    = get_input("目标文件夹路径", default_dst)
            copy_files_from_excel(excel_path, src_dir, dst_dir)

        elif choice == '3':
            print("\n--- 重命名文件 ---")
            excel_path = get_input("Excel 文件路径", default_excel)
            src_dir    = get_input("源文件夹路径", default_src)
            rename_files_from_excel(excel_path, src_dir)

        elif choice == '4':
            print("\n--- 记录写入 ---")
            excel_path = get_input("Excel 文件路径", default_excel)
            src_dir    = get_input("源文件夹路径", default_src)
            inc_sub    = get_input("是否包含子文件夹？(y/n)", "n").strip().lower() == 'y'
            write_filenames_to_excel(excel_path, src_dir, inc_sub)

        elif choice == '5':
            print("\n--- 记录删除 ---")
            excel_path = get_input("Excel 文件路径", default_excel)
            src_dir    = get_input("源文件夹路径", default_src)
            inc_sub    = get_input("是否包含子文件夹？(y/n)", "n").strip().lower() == 'y'
            match_field = get_input("匹配的字段名（原文件名/现文件名）", "原文件名")
            # 简单修正可能的误输入，默认仍为“原文件名”
            if match_field not in ("原文件名", "现文件名"):
                print(f"警告：未知字段‘{match_field}’，将使用默认的‘原文件名’。")
                match_field = "原文件名"
            delete_records_from_excel(excel_path, src_dir, inc_sub, match_field)

        else:
            print("无效的选择，请输入 0-5 之间的数字。")

        input("\n按回车键继续...")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n操作被用户中断。")
    except Exception as e:
        print(f"发生未预期的错误: {e}")